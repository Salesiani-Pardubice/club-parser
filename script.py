import openpyxl
import re


def main():
    # Open the file
    path = "input.xlsx"  # please specific file path
    using_col = [6, 7, 8, 14, 15, 16, 17, 18, 19, 20, 21, 25, 28]

    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    # Get the working column
    working_col = 0

    for col in range(1, sheet.max_column+1):
        if (sheet.cell(row=1, column=col).value == "Kroužky"):
            working_col = col
            break

    # print("The working coulm is: ", working_col)

    # Create a camp collection
    camps = []
    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row=row, column=working_col).value
        if (cell != None):
            cell_camps = cell.split(";")
            for value in cell_camps:
                if value not in camps and value != "":
                    camps.append(value)
    # print(camps)
    for camp in camps:
        # Create separe file for every camp
        wb_new = openpyxl.Workbook()
        sheet_new = wb_new.active
        # Copy header row
        new_col = 1
        for col in range(1, sheet.max_column + 1):
            if col in using_col:
                sheet_new.cell(row=1, column=new_col).value = sheet.cell(
                    row=1, column=col).value
                new_col += 1
        # Copy values
        new_row = 2
        for row in range(2, sheet.max_row + 1):
            # check if row contain current camp
            if sheet.cell(row=row, column=working_col).value != None and camp in sheet.cell(row=row, column=working_col).value:
                new_col = 1
                for col in range(1, sheet.max_column + 1):
                    if col in using_col:
                        sheet_new.cell(row=new_row, column=new_col).value = sheet.cell(
                            row=row, column=col).value
                        new_col += 1
                new_row += 1
        filename = "" + re.sub(r'\([^)]*\)', '', camp) + ".xlsx"
        # wb_new.save("output.xlsx")
        wb_new.save(filename)


if __name__ == "__main__":
    main()
